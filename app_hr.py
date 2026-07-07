"""
干部教育培训统计 - 工程级 Streamlit 应用
"""

import streamlit as st
import pandas as pd
import numpy as np
import os, re, io
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
from collections import defaultdict
import json as _json
from hr_analyzer.llm import ask_ai_recommendation, parse_recommendation_json

st.set_page_config(page_title="干部教育培训统计", layout="wide")

# ============================================================
# Config
# ============================================================

FIELD_ALIASES: Dict[str, List[str]] = {
    "employee_code": ["集团员工编码", "员工编码", "工号", "人员编码", "编码"],
    "name": ["人员姓名", "姓名", "名字", "员工姓名"],
    "cadre_level": ["干部标识", "干部级别", "职级", "级别"],
    "unit": ["单位名称", "单位", "所属单位"],
    "department": ["部门名称", "部门", "所属部门"],
    "position": ["职务", "职位", "岗位"],
    "training_method": ["培训方式", "学习方式"],
    "hours": ["学时", "培训学时"],
    "start_date": ["开始学习时间", "开始时间"],
    "end_date": ["完成学习时间", "结束时间"],
    "course_name": ["来源信息", "培训班次名称", "课程名称", "培训项目"],
    "organizer": ["主办单位"],
    "institution": ["培训机构"],
}

VALID_TRAINING_METHODS = [
    "党委(党组)理论学习中心组学习",
    "脱产培训(3天以上)",
    "集中宣讲/专题讲座",
]

RANGE_TRAINING_METHODS = ["脱产培训(3天以上)", "集中宣讲/专题讲座"]

HOURS_RANGE_40_90 = (40, 90)
HOURS_RANGE_90_PLUS = 90

# ============================================================
# Utils
# ============================================================

def resolve_field(df: pd.DataFrame, standard_name: str) -> Optional[str]:
    aliases = FIELD_ALIASES.get(standard_name, [])
    df_cols_lower = {c.lower().strip(): c for c in df.columns}
    for alias in aliases:
        key = alias.lower().strip()
        if key in df_cols_lower:
            return df_cols_lower[key]
    return None


@st.cache_data(show_spinner=False)
def safe_read_excel(file) -> pd.DataFrame:
    if hasattr(file, "name") and file.name.endswith(".csv"):
        return pd.read_csv(file)
    xl = pd.ExcelFile(file)
    if len(xl.sheet_names) == 1:
        return pd.read_excel(file)
    largest = max(xl.sheet_names, key=lambda s: xl.parse(s).shape[0])
    return pd.read_excel(file, sheet_name=largest)


def safe_parse_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def safe_parse_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


# ============================================================
# Validator
# ============================================================

@dataclass
class QualityReport:
    person_rows: int = 0
    training_rows: int = 0
    issues: List[Dict[str, Any]] = field(default_factory=list)
    is_valid: bool = True

    @property
    def issue_count(self) -> int:
        return len(self.issues)

    def to_dataframe(self) -> pd.DataFrame:
        if not self.issues:
            return pd.DataFrame({"状态": ["ok"]})
        return pd.DataFrame(self.issues)


def _check_missing(df: pd.DataFrame, label: str, col: str) -> List[Dict[str, Any]]:
    mask = df[col].isna() | (df[col].astype(str).str.strip() == "")
    if mask.any():
        return [{"表": label, "问题类型": "缺失值", "字段": col, "数量": int(mask.sum()), "详情": f"{mask.sum()} 行 {col} 为空"}]
    return []


def _check_dupes(df: pd.DataFrame, label: str, col: str) -> List[Dict[str, Any]]:
    valid = df[col].dropna()
    dup_count = valid.duplicated().sum()
    if dup_count > 0:
        return [{"表": label, "问题类型": "重复编码", "字段": col, "数量": dup_count, "详情": f"{dup_count} 条重复"}]
    return []


def _check_bad_dates(df: pd.DataFrame, label: str, col: str) -> List[Dict[str, Any]]:
    parsed = safe_parse_date(df[col])
    invalid = df[col].notna() & parsed.isna()
    if invalid.any():
        return [{"表": label, "问题类型": "非法日期", "字段": col, "数量": int(invalid.sum()), "详情": f"{invalid.sum()} 行无法解析"}]
    return []


def _check_bad_hours(df: pd.DataFrame, label: str, col: str) -> List[Dict[str, Any]]:
    parsed = safe_parse_numeric(df[col])
    issues = []
    invalid = df[col].notna() & parsed.isna()
    negative = pd.notna(parsed) & (parsed < 0)
    if invalid.any():
        issues.append({"表": label, "问题类型": "非法学时", "字段": col, "数量": int(invalid.sum()), "详情": f"{invalid.sum()} 行非数值"})
    if negative.any():
        issues.append({"表": label, "问题类型": "负数学时", "字段": col, "数量": int(negative.sum()), "详情": f"{negative.sum()} 行为负数"})
    return issues


@st.cache_data(show_spinner=False)
def run_validation(
    person_df: pd.DataFrame,
    training_df: pd.DataFrame,
    e1: str, e2: str, h: str, s: str, ed: str, m: str,
) -> QualityReport:
    report = QualityReport(person_rows=len(person_df), training_rows=len(training_df))
    issues = []
    issues += _check_missing(person_df, "表1", e1)
    issues += _check_dupes(person_df, "表1", e1)
    issues += _check_missing(training_df, "表2", e2)
    issues += _check_bad_hours(training_df, "表2", h)
    issues += _check_bad_dates(training_df, "表2", s)
    issues += _check_bad_dates(training_df, "表2", ed)
    # field missing warnings
    for std_name in ["name", "course_name", "organizer", "institution"]:
        table = "表1" if std_name in ("name",) else "表2"
        df_ref = person_df if table == "表1" else training_df
        if resolve_field(df_ref, std_name) is None:
            issues.append({"表": table, "问题类型": "字段缺失", "字段": std_name, "数量": 0, "详情": f"未找到 {std_name}"})
    report.issues = issues
    report.is_valid = len([i for i in issues if i["问题类型"] in ("缺失值","重复编码","非法日期","非法学时")]) == 0
    return report


# ============================================================
# Cleaner
# ============================================================

@st.cache_data(show_spinner=False)
def clean_person_table(df: pd.DataFrame, col_employee: str) -> pd.DataFrame:
    df = df.copy()
    df = df[df[col_employee].notna() & (df[col_employee].astype(str).str.strip() != "")].copy()
    df = df.drop_duplicates(subset=[col_employee], keep="first").copy()
    df[col_employee] = df[col_employee].astype(str).str.strip()
    return df


@st.cache_data(show_spinner=False)
def clean_training_table(df: pd.DataFrame, e: str, h: str, s: str, ed: str) -> pd.DataFrame:
    df = df.copy()
    df = df[df[e].notna() & (df[e].astype(str).str.strip() != "")].copy()
    df[e] = df[e].astype(str).str.strip()
    df[s] = safe_parse_date(df[s])
    df[ed] = safe_parse_date(df[ed])
    df[h] = safe_parse_numeric(df[h])
    df = df.dropna(subset=[h]).copy()
    return df


# ============================================================
# Analyzer
# ============================================================

@dataclass
class AnalysisResult:
    merged_count: int = 0
    filtered_count: int = 0
    total_hours: Optional[pd.DataFrame] = None
    recent_40_90: Optional[pd.DataFrame] = None
    recent_90_plus: Optional[pd.DataFrame] = None
    report_df: Optional[pd.DataFrame] = None
    person_count: int = 0
    total_hours_sum: float = 0.0
    count_40_90: int = 0
    count_90_plus: int = 0


def merge_tables(person_df, training_df, e1, e2):
    merged = person_df[[e1]].merge(training_df, left_on=e1, right_on=e2, how="inner").copy()
    merged.rename(columns={e1: "employee_code"}, inplace=True)
    return merged


def filter_training_methods(df, col_method):
    return df[df[col_method].isin(VALID_TRAINING_METHODS)].copy()


def calculate_total_hours(df, col_hours):
    total = df.groupby("employee_code")[col_hours].sum().reset_index()
    total.columns = ["employee_code", "累计培训学时"]
    total["累计培训学时"] = total["累计培训学时"].round(1)
    return total


def _effective_date(row, col_start, col_end):
    if pd.notna(row.get(col_end)):
        return row[col_end]
    return row[col_start]


def _extract_range(filtered_df, col_hours, col_start, col_end, col_method, hours_low, hours_high):
    df = filtered_df.copy()
    df = df[df[col_method].isin(RANGE_TRAINING_METHODS)].copy()
    df = df[df[col_hours] >= hours_low].copy()
    if hours_high is not None:
        df = df[df[col_hours] < hours_high].copy()
    if df.empty:
        return pd.DataFrame()

    df["_sort_date"] = df.apply(lambda r: _effective_date(r, col_start, col_end), axis=1)
    df = df.dropna(subset=["_sort_date"]).copy()
    if df.empty:
        return pd.DataFrame()

    latest = df.sort_values("_sort_date", ascending=False).groupby("employee_code").first().reset_index()

    detail_cols = ["employee_code", "来源信息", "开始学习时间", "完成学习时间", "培训机构", "主办单位", col_hours, col_method]
    available = [c for c in detail_cols if c in latest.columns]
    result = latest[available].copy()

    for dt_col in ["开始学习时间", "完成学习时间"]:
        if dt_col in result.columns:
            result[dt_col] = result[dt_col].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else "")

    result = result.drop(columns=["_sort_date"], errors="ignore")
    return result


def extract_range_40_90(filtered_df, col_hours, col_start, col_end, col_method):
    return _extract_range(filtered_df, col_hours, col_start, col_end, col_method, HOURS_RANGE_40_90[0], HOURS_RANGE_40_90[1])


def extract_range_90_plus(filtered_df, col_hours, col_start, col_end, col_method):
    return _extract_range(filtered_df, col_hours, col_start, col_end, col_method, HOURS_RANGE_90_PLUS, None)


def _attach_range(report, range_df, prefix):
    if range_df.empty:
        report[f"{prefix}·班次名称（表2）"] = "无符合条件记录"
        report[f"{prefix}·开始学习时间（表2）"] = ""
        report[f"{prefix}·完成学习时间（表2）"] = ""
        report[f"{prefix}·培训机构（表2）"] = ""
        report[f"{prefix}·主办单位（表2）"] = ""
        return report
    details = range_df.copy()
    col_map = {}
    for src_col in details.columns:
        if src_col == "employee_code":
            continue
        col_map[src_col] = f"{prefix}·{src_col}（表2）"
    details = details.rename(columns=col_map)
    report = report.merge(details, on="employee_code", how="left")
    for col in col_map.values():
        if col in report.columns:
            report[col] = report[col].fillna("无符合条件记录")
    return report


def build_integrated_report(person_df, e1, total_hours, recent_40_90, recent_90_plus):
    info_cols = [e1]
    for std_name in ["name", "cadre_level", "unit", "department"]:
        resolved = resolve_field(person_df, std_name)
        if resolved and resolved not in info_cols:
            info_cols.append(resolved)

    info_df = person_df[info_cols].copy()
    info_df = info_df.rename(columns={e1: "employee_code"})

    report = info_df.merge(total_hours, on="employee_code", how="left")
    report["累计培训学时"] = report["累计培训学时"].fillna(0.0)

    report = _attach_range(report, recent_40_90, "40~90")
    report = _attach_range(report, recent_90_plus, "90+")

    report["累计培训学时"] = pd.to_numeric(report["累计培训学时"], errors="coerce")

    rename_map = {}
    for col in report.columns:
        if col == "employee_code":
            rename_map[col] = "员工编码（表1）"
        elif col == "累计培训学时":
            continue
        elif resolve_field(person_df, "name") and col == resolve_field(person_df, "name"):
            rename_map[col] = "姓名（表1）"
        elif resolve_field(person_df, "cadre_level") and col == resolve_field(person_df, "cadre_level"):
            rename_map[col] = "干部标识（表1）"
        elif resolve_field(person_df, "unit") and col == resolve_field(person_df, "unit"):
            rename_map[col] = "单位（表1）"
        elif resolve_field(person_df, "department") and col == resolve_field(person_df, "department"):
            rename_map[col] = "部门（表1）"
    report = report.rename(columns=rename_map)
    return report


def run_analysis(person_df, training_df, e1, e2, h, s, ed, m):
    result = AnalysisResult()
    result.person_count = len(person_df)
    merged = merge_tables(person_df, training_df, e1, e2)
    result.merged_count = len(merged)
    filtered = filter_training_methods(merged, m)
    result.filtered_count = len(filtered)
    if filtered.empty:
        return result
    result.total_hours = calculate_total_hours(filtered, h)
    if result.total_hours is not None:
        result.total_hours_sum = result.total_hours["累计培训学时"].sum()
    result.recent_40_90 = extract_range_40_90(filtered, h, s, ed, m)
    result.count_40_90 = len(result.recent_40_90) if result.recent_40_90 is not None else 0
    result.recent_90_plus = extract_range_90_plus(filtered, h, s, ed, m)
    result.count_90_plus = len(result.recent_90_plus) if result.recent_90_plus is not None else 0
    result.report_df = build_integrated_report(person_df, e1, result.total_hours, result.recent_40_90, result.recent_90_plus)
    return result


# ============================================================
# CandidateBuilder — 培训推荐排序
# ============================================================

@dataclass
class CandidateProfile:
    """候选人画像 — 用于 AI 推荐排序。"""
    employee_code: str = ""
    name: str = ""
    cadre_level: str = ""
    unit: str = ""
    department: str = ""
    position: str = ""
    total_hours: float = 0.0
    recent_trainings: list = field(default_factory=list)
    last_training: Optional[dict] = None


def build_candidate_profiles(
    person_df, training_df, total_hours,
    e1, e2, m, s, ed,
) -> Dict[str, CandidateProfile]:
    """构建候选人画像字典。"""
    profiles: Dict[str, CandidateProfile] = {}

    name_col = resolve_field(person_df, "name")
    cadre_col = resolve_field(person_df, "cadre_level")
    unit_col = resolve_field(person_df, "unit")
    dept_col = resolve_field(person_df, "department")
    pos_col = resolve_field(person_df, "position")

    hours_map: Dict[str, float] = {}
    if total_hours is not None and not total_hours.empty:
        for _, row in total_hours.iterrows():
            hours_map[str(row["employee_code"]).strip()] = float(row.get("累计培训学时", 0))

    train_by_person: defaultdict = defaultdict(list)
    if training_df is not None and not training_df.empty:
        course_col = resolve_field(training_df, "course_name")
        org_col = resolve_field(training_df, "institution")
        host_col = resolve_field(training_df, "organizer")

        for _, row in training_df.iterrows():
            code = str(row[e2]).strip()
            method_val = row.get(m)
            method = str(method_val) if pd.notna(method_val) else ""
            if method not in VALID_TRAINING_METHODS:
                continue
            end_dt = row.get(ed)
            start_dt = row.get(s)
            sort_date = end_dt if pd.notna(end_dt) else start_dt

            train_by_person[code].append({
                "课程": str(row.get(course_col, "")) if course_col and pd.notna(row.get(course_col, "")) else "",
                "培训方式": method,
                "开始时间": start_dt.strftime("%Y-%m-%d") if pd.notna(start_dt) else "",
                "结束时间": end_dt.strftime("%Y-%m-%d") if pd.notna(end_dt) else "",
                "培训机构": str(row.get(org_col, "")) if org_col and pd.notna(row.get(org_col, "")) else "",
                "主办单位": str(row.get(host_col, "")) if host_col and pd.notna(row.get(host_col, "")) else "",
                "_sort_date": sort_date,
            })

    for _, row in person_df.iterrows():
        code = str(row[e1]).strip()
        profile = CandidateProfile(
            employee_code=code,
            name=str(row.get(name_col, "")) if name_col and pd.notna(row.get(name_col, "")) else "",
            cadre_level=str(row.get(cadre_col, "")) if cadre_col and pd.notna(row.get(cadre_col, "")) else "",
            unit=str(row.get(unit_col, "")) if unit_col and pd.notna(row.get(unit_col, "")) else "",
            department=str(row.get(dept_col, "")) if dept_col and pd.notna(row.get(dept_col, "")) else "",
            position=str(row.get(pos_col, "")) if pos_col and pd.notna(row.get(pos_col, "")) else "",
            total_hours=hours_map.get(code, 0.0),
        )

        trainings = train_by_person.get(code, [])
        trainings_sorted = sorted(
            trainings,
            key=lambda t: t["_sort_date"] if pd.notna(t["_sort_date"]) else pd.Timestamp.min,
            reverse=True,
        )

        seen: set = set()
        unique = []
        for t in trainings_sorted:
            course = t["课程"]
            if course and course not in seen:
                seen.add(course)
                unique.append(t)

        profile.recent_trainings = unique[:20]
        profile.last_training = unique[0] if unique else None
        profiles[code] = profile

    return profiles


def recommend_candidates(
    profiles: Dict[str, CandidateProfile],
    training_name: str = "",
    training_goal: str = "",
    cadre_level_filter: Optional[List[str]] = None,
    exclude_training_name: bool = True,
    exclude_recent_similar: bool = True,
    recent_years: int = 5,
    min_hours: float = 0,
    max_hours: float = float("inf"),
    max_candidates: int = 200,
) -> List[CandidateProfile]:
    """硬规则筛选推荐候选人（纯 Python/Pandas，AI 不参与）。

    筛选规则：
    1. 学时范围过滤
    2. 干部级别过滤
    3. 排除已参加同名培训
    4. 排除近 N 年参加过类似培训（脱产/集中宣讲）>=3 次
    5. 按累计学时升序（学时少的优先推荐）
    """
    candidates: List[CandidateProfile] = []
    cutoff = pd.Timestamp.now() - pd.DateOffset(years=recent_years)

    for code, profile in profiles.items():
        # 1. 学时过滤
        if profile.total_hours < min_hours or profile.total_hours > max_hours:
            continue

        # 2. 干部级别过滤
        if cadre_level_filter and profile.cadre_level not in cadre_level_filter:
            continue

        # 3. 排除已参加同名培训
        if exclude_training_name and training_name:
            attended = any(
                training_name in t.get("课程", "") for t in profile.recent_trainings
            )
            if attended:
                continue

        # 4. 排除近 N 年类似培训过多（脱产培训 或 集中宣讲 >= 3 次）
        if exclude_recent_similar:
            recent_similar = [
                t for t in profile.recent_trainings
                if pd.notna(t["_sort_date"]) and t["_sort_date"] >= cutoff
                and ("脱产培训" in t.get("培训方式", "")
                     or "集中宣讲" in t.get("培训方式", ""))
            ]
            if len(recent_similar) >= 3:
                continue

        candidates.append(profile)

    # 5. 按累计学时升序（学时少的优先推荐）
    candidates.sort(key=lambda p: p.total_hours)
    return candidates[:max_candidates]


# ============================================================
# Exporter
# ============================================================

@st.cache_data(show_spinner=False)
def export_excel(report_df, quality_report, analysis_result) -> BytesIO:
    """Export Excel with openpyxl formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "整合报表"

    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for ci, cn in enumerate(report_df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(cn))
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin_border

    for ri, row in enumerate(report_df.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_w = 0
        for cell in col_cells[:200]:
            if cell.value is None:
                continue
            w = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
            max_w = max(max_w, w)
        ws.column_dimensions[col_letter].width = max(min(max_w + 4, 45), 8)

    ws.freeze_panes = "A2"
    if len(report_df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(report_df.columns))}{len(report_df)+1}"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_csv(report_df: pd.DataFrame) -> bytes:
    return report_df.to_csv(index=False).encode("utf-8-sig")


@st.cache_data(show_spinner=False)
def build_markdown_report(report_df: pd.DataFrame, analysis_result: AnalysisResult) -> str:
    lines = []

    # Module 1
    lines.append("## 模块一：人员基本信息")
    lines.append("")
    person_cols = [c for c in report_df.columns if "（表1）" in c]
    if "累计培训学时" in report_df.columns:
        person_cols.append("累计培训学时")
    if person_cols:
        lines.append(report_df[person_cols].head(20).to_markdown(index=False))
    lines.append(f"*（共 {len(report_df)} 人，仅展示前 20 行）*")
    lines.append("")
    lines.append("---")

    # Module 2
    lines.append("## 模块二：累计培训学时统计")
    lines.append("")
    lines.append(f"- 总人数：{analysis_result.person_count}")
    lines.append(f"- 有效培训记录：{analysis_result.filtered_count} 条")
    lines.append(f"- 累计总学时：**{analysis_result.total_hours_sum:,.0f}**")
    if analysis_result.total_hours is not None:
        th = analysis_result.total_hours
        lines.append(f"- 人均学时：{th['累计培训学时'].mean():.1f}")
        lines.append(f"- 最高学时：{th['累计培训学时'].max():.0f}")
        lines.append(f"- 学时中位数：{th['累计培训学时'].median():.1f}")
    lines.append("")
    lines.append("---")

    # Module 3
    lines.append("## 模块三：40~90 学时区间（脱产培训 / 集中宣讲）")
    lines.append("")
    lines.append(f"- 符合条件：{analysis_result.count_40_90} 人")
    range_cols = [c for c in report_df.columns if "40~90" in c]
    if range_cols:
        col = next((c for c in range_cols if "班次名称" in c or "来源信息" in c), range_cols[0])
        valid = report_df[report_df[col] != "无符合条件记录"]
        if not valid.empty:
            lines.append(valid[range_cols].head(20).to_markdown(index=False))
            lines.append(f"*（共 {len(valid)} 人，仅展示前 20 行）*")
    lines.append("")
    lines.append("---")

    # Module 4
    lines.append("## 模块四：90 学时以上区间（脱产培训 / 集中宣讲）")
    lines.append("")
    lines.append(f"- 符合条件：{analysis_result.count_90_plus} 人")
    range_cols = [c for c in report_df.columns if "90+" in c]
    if range_cols:
        col = next((c for c in range_cols if "班次名称" in c or "来源信息" in c), range_cols[0])
        valid = report_df[report_df[col] != "无符合条件记录"]
        if not valid.empty:
            lines.append(valid[range_cols].head(20).to_markdown(index=False))
            lines.append(f"*（共 {len(valid)} 人，仅展示前 20 行）*")

    return "\n".join(lines)




# ============================================================
# UI
# ============================================================

if "analysis_done" not in st.session_state:
    st.session_state.analysis_done = False
if "analysis_result" not in st.session_state:
    st.session_state.analysis_result = None
if "quality_report" not in st.session_state:
    st.session_state.quality_report = None
if "person_df" not in st.session_state:
    st.session_state.person_df = None
if "field_mapping" not in st.session_state:
    st.session_state.field_mapping = {}
if "rec_candidates" not in st.session_state:
    st.session_state.rec_candidates = []
if "rec_done" not in st.session_state:
    st.session_state.rec_done = False


def _build_field_mapping(person_df, training_df):
    mapping = {}
    for std_name in FIELD_ALIASES:
        col = resolve_field(training_df, std_name)
        if col is None:
            col = resolve_field(person_df, std_name)
        if col:
            mapping[std_name] = col
    return mapping


st.title("干部教育培训统计")

st.markdown("### 表1：干部基础信息")
file1 = st.file_uploader("上传干部基础信息", type=["csv", "xlsx"], key="f1")
st.markdown("### 表2：培训学时记录")
file2 = st.file_uploader("上传培训学时记录", type=["csv", "xlsx"], key="f2")

if file1 and file2:
    try:
        df_person_raw = safe_read_excel(file1)
        df_training_raw = safe_read_excel(file2)
    except Exception as exc:
        st.error(f"文件读取失败：{exc}")
        st.stop()

    field_map = _build_field_mapping(df_person_raw, df_training_raw)
    st.session_state.field_mapping = field_map

    with st.expander("🔍 字段自动识别结果", expanded=False):
        for std, actual in field_map.items():
            st.caption(f"`{std}` → **{actual}**")
        missing = [s for s in ["employee_code", "hours", "training_method"] if s not in field_map]
        if missing:
            st.warning(f"⚠️ 未识别字段：{', '.join(missing)}")

    required = ["employee_code", "hours", "training_method"]
    if not all(k in field_map for k in required):
        st.error(f"缺少必需字段：{[k for k in required if k not in field_map]}")
        st.stop()

    e1 = field_map["employee_code"]
    e2 = field_map["employee_code"]
    h_col = field_map["hours"]
    s_col = field_map.get("start_date", "")
    ed_col = field_map.get("end_date", "")
    m_col = field_map["training_method"]

    df_person = clean_person_table(df_person_raw, e1)
    df_training = clean_training_table(df_training_raw, e2, h_col, s_col, ed_col)

    qr = run_validation(df_person, df_training, e1, e2, h_col, s_col, ed_col, m_col)
    st.session_state.quality_report = qr
    st.session_state.person_df = df_person

    with st.expander(f"📋 数据质量报告（{qr.issue_count} 项）", expanded=qr.issue_count > 0):
        if qr.issue_count == 0:
            st.success("✅ 数据质量检查通过")
        else:
            st.dataframe(qr.to_dataframe(), use_container_width=True)

    st.subheader("数据预览")
    c1, c2 = st.columns(2)
    c1.metric("表1：干部基础信息", f"{len(df_person)} 人")
    c2.metric("表2：培训学时记录", f"{len(df_training)} 条")

    # --- 统计分析按钮（仅执行分析，不渲染结果） ---
    if st.button("🚀 开始统计分析", type="primary", use_container_width=True):
        with st.spinner("正在执行统计分析..."):
            try:
                result = run_analysis(df_person, df_training, e1, e2, h_col, s_col, ed_col, m_col)
                st.session_state.analysis_result = result
                st.session_state.analysis_done = True
            except Exception as exc:
                st.error(f"分析执行失败：{exc}")
                st.stop()

    # ================================================================
    # 以下全部在按钮块外部，用 session_state 守卫，不会因重跑消失
    # ================================================================

    if st.session_state.get("analysis_done"):
        result = st.session_state.analysis_result
        df_person = st.session_state.person_df
        qr = st.session_state.quality_report

        st.subheader("📊 统计概览")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("干部总数", result.person_count)
        m2.metric("有效培训记录", result.filtered_count)
        m3.metric("累计总学时", f"{result.total_hours_sum:,.0f}")
        m4.metric("人均学时", f"{result.total_hours_sum / max(result.person_count, 1):.1f}")

        c1, c2 = st.columns(2)
        c1.metric("40~90学时人数", result.count_40_90)
        c2.metric("90学时以上人数", result.count_90_plus)

        # Markdown 报表
        with st.expander("📝 四模块报表（Markdown）", expanded=False):
            md_report = build_markdown_report(result.report_df, result)
            st.markdown(md_report)

        # 导出（在按钮块外部，可以正常点击）
        col_xl, col_csv = st.columns(2)
        with col_xl:
            excel_bytes = export_excel(result.report_df, qr, result)
            st.download_button(
                "📥 下载 Excel 报告", excel_bytes, "干部教育培训统计.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_csv:
            csv_bytes = export_csv(result.report_df)
            st.download_button(
                "📥 下载 CSV", csv_bytes, "干部教育培训统计.csv",
                mime="text/csv", use_container_width=True,
            )

        # 培训推荐排序
        st.divider()
        st.subheader("🎯 培训推荐排序")

        with st.expander("📝 培训班信息", expanded=True):
            rec_name = st.text_input(
                "培训班名称",
                placeholder="例如：新时代年轻干部能力提升班",
                key="rec_name",
            )
            rec_goal = st.text_area(
                "培训目标",
                placeholder="例如：培养35~45岁年轻干部，提高战略思维、数字化能力、党建能力",
                key="rec_goal",
            )
            ca, cb = st.columns(2)
            with ca:
                rec_min_h = st.number_input("最低累计学时", min_value=0.0, value=0.0, step=1.0, key="rec_min_h")
                rec_max_n = st.number_input("最大候选人数", min_value=10, max_value=500, value=200, step=10, key="rec_max_n")
            with cb:
                rec_max_h = st.number_input("最高累计学时（0=不限）", min_value=0.0, value=0.0, step=1.0, key="rec_max_h")
                rec_excl = st.checkbox("排除已参加同名培训的人员", value=True, key="rec_excl")

            if st.button("🔍 执行硬规则筛选", key="rec_filter_btn", use_container_width=True):
                with st.spinner("正在构建候选人画像并执行硬规则筛选..."):
                    try:
                        profiles = build_candidate_profiles(
                            df_person, df_training,
                            result.total_hours,
                            e1, e2, m_col, s_col, ed_col,
                        )
                        max_h = rec_max_h if rec_max_h > 0 else float("inf")
                        candidates = recommend_candidates(
                            profiles, training_name=rec_name, training_goal=rec_goal,
                            exclude_training_name=rec_excl, min_hours=rec_min_h,
                            max_hours=max_h, max_candidates=rec_max_n,
                        )
                        st.session_state.rec_candidates = candidates
                        st.session_state.rec_done = True
                    except Exception as exc:
                        st.error(f"候选人筛选失败：{exc}")

        if st.session_state.get("rec_done"):
            candidates = st.session_state.rec_candidates
            st.success(f"✅ 硬规则筛选完成：{len(candidates)} 名候选人")

            with st.expander(f"📋 候选人列表（{len(candidates)} 人）", expanded=False):
                cand_data = []
                for c in candidates:
                    last = c.last_training["课程"] if c.last_training else "无"
                    cand_data.append({
                        "姓名": c.name, "干部级别": c.cadre_level,
                        "单位": c.unit, "部门": c.department,
                        "职务": c.position, "累计学时": c.total_hours,
                        "最近培训": last,
                    })
                st.dataframe(pd.DataFrame(cand_data), use_container_width=True, height=400)
                st.caption(f"共 {len(candidates)} 人 | 按累计学时升序排列")

            with st.expander("⚙️ DeepSeek API 配置（推荐排序）", expanded=False):
                rec_api_key = st.text_input("API Key", type="password", placeholder="sk-...", key="rec_ds_key")

            if rec_api_key:
                if st.button("🤖 AI 推荐排序（JSON）", key="rec_ai_btn", use_container_width=True):
                    if not rec_name:
                        st.warning("请先填写培训班名称")
                    else:
                        with st.spinner(f"AI 正在对 {len(candidates)} 名候选人进行匹配度评估..."):
                            try:
                                from openai import OpenAI
                                ai_client = OpenAI(api_key=rec_api_key, base_url="https://api.deepseek.com")
                                answer = ask_ai_recommendation(ai_client, candidates, rec_name, rec_goal)
                                rec_df = parse_recommendation_json(answer)

                                if "error" in rec_df.columns:
                                    st.error("AI 返回格式异常，请重试")
                                    st.text(answer[:1000])
                                else:
                                    st.session_state.rec_df = rec_df
                            except Exception as exc:
                                st.error(f"AI 推荐排序失败：{exc}")

            if "rec_df" in st.session_state and st.session_state.rec_df is not None:
                rec_df = st.session_state.rec_df
                st.success(f"AI 推荐排序完成：{len(rec_df)} 人已评分")
                st.markdown("### AI 推荐排序结果")

                ca2, cb2, cc2 = st.columns(3)
                ca2.metric("平均分", f"{rec_df['score'].mean():.0f}")
                cb2.metric("最高分", f"{rec_df['score'].max():.0f}")
                cc2.metric("推荐人数（>=80）", len(rec_df[rec_df['score'] >= 80]))

                display_df = rec_df.rename(columns={
                    "employee_code": "员工编码",
                    "name": "姓名",
                    "score": "评分",
                    "level": "匹配度",
                    "reason": "推荐理由",
                    "risk": "不推荐理由",
                    "priority": "优先安排",
                    "follow_up": "后续培养",
                    "backup": "替补建议",
                })
                st.dataframe(
                    display_df, use_container_width=True, height=500,
                    column_config={
                        "评分": st.column_config.ProgressColumn(
                            "评分", min_value=0, max_value=100, format="%d"
                        ),
                    },
                    hide_index=True,
                )

                col_xl2, col_csv2 = st.columns(2)
                with col_xl2:
                    buf = io.BytesIO()
                    rec_df.to_excel(buf, index=False, engine="openpyxl")
                    st.download_button(
                        "📥 下载推荐结果 Excel", buf.getvalue(),
                        "培训推荐排序结果.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="rec_xl_dl", use_container_width=True,
                    )
                with col_csv2:
                    st.download_button(
                        "📥 下载推荐结果 CSV",
                        rec_df.to_csv(index=False).encode("utf-8-sig"),
                        "培训推荐排序结果.csv", mime="text/csv",
                        key="rec_csv_dl", use_container_width=True,
                    )


# ============================================================
# Self-check
# ============================================================

if __name__ == "__main__":
    print("Config: FIELD_ALIASES =", len(FIELD_ALIASES))
    print("Utils: resolve_field, safe_read_excel")
    print("Validator: QualityReport, run_validation")
    print("Cleaner: clean_person_table, clean_training_table")
    print("Analyzer: merge_tables, filter_training_methods, calculate_total_hours")
    print("CandidateBuilder: CandidateProfile, build_candidate_profiles, recommend_candidates")
    print("Exporter: export_excel, export_csv, build_markdown_report")
    print("LLM: ask_ai_recommendation, parse_recommendation_json (from hr_analyzer.llm)")
    print("UI: Streamlit interface")
    print("OK - All modules present")
